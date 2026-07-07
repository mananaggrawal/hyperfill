"""
Excel builder for RFP annexures, BOQs, and pricing sheets.

Many commercial RFPs require an Excel enclosure (a pre-bid query format, a
Bill of Quantities, a price schedule) rather than — or in addition to — a
Word annexure. This module covers the two patterns that actually come up:

  1. fill_template() — the RFP ships its own .xlsx template (e.g. a pre-bid
     query format, a technical-scoring rubric). Open it, drop values into
     named cells, save a copy. This preserves the buyer's own formatting,
     column widths, and any formulas already in the sheet — don't rebuild
     their template from scratch.

  2. build_workbook() — no template exists (e.g. a BOQ or price break-up you
     have to construct yourself). Build a simple styled workbook: bold header
     row, borders, company details in a footer.

Requires `openpyxl` (installed silently: `pip install openpyxl --break-system-packages`).
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import bidder_profile as profile

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")


def fill_template(template_path, output_path, values: dict, sheet_name: str = None) -> str:
    """Fill cells in a copy of an existing .xlsx template.

    `values` maps cell references to content, e.g. {"B4": "Acme Technology
    Private Limited", "B5": "U72900MH2022PTC384814"}. Pass a dict of dicts
    keyed by sheet name if the template has multiple sheets to fill:
    {"Sheet1": {"B4": "..."}, "Pricing": {"C10": 1500}}.

    Never touches formulas or formatting already in the template — it only
    sets cell values. Returns the output path.
    """
    template_path = str(template_path)
    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)

    # Detect whether `values` is single-sheet ({"B4": ...}) or multi-sheet
    # ({"Sheet1": {"B4": ...}}) by checking if all values are dicts.
    is_multi_sheet = values and all(isinstance(v, dict) for v in values.values())

    if is_multi_sheet:
        for sheet, cells in values.items():
            if sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{sheet}' not found in {template_path}. "
                                f"Available: {wb.sheetnames}")
            ws = wb[sheet]
            for ref, val in cells.items():
                ws[ref] = val
    else:
        ws = wb[sheet_name] if sheet_name else wb.active
        for ref, val in values.items():
            ws[ref] = val

    wb.save(output_path)
    return output_path


def build_workbook(title: str, headers: list, rows: list, output_path,
                    col_widths: list = None, sign_off: bool = True) -> str:
    """Build a new styled workbook from scratch — for BOQs, price schedules,
    or any tabular annexure the RFP doesn't supply a template for.

    `headers` is a list of column titles. `rows` is a list of row tuples,
    same length as `headers`. `col_widths` is optional (character widths).
    If `sign_off` is True, appends the authorised signatory's name and
    designation a couple of rows below the table.
    """
    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = 3
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row_values in enumerate(rows, start=header_row + 1):
        for col, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    if col_widths:
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    else:
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    if sign_off:
        sign_row = header_row + len(rows) + 3
        s = profile.AUTHORISED_SIGNATORY
        ws.cell(row=sign_row, column=1,
                value=f"For {profile.COMPANY.get('legal_name', '')}").font = Font(bold=True)
        ws.cell(row=sign_row + 2, column=1,
                value=f"Name: {s.get('name', '')}").font = Font(bold=True)
        ws.cell(row=sign_row + 3, column=1,
                value=f"Designation: {s.get('designation', '')}").font = Font(bold=True)

    wb.save(output_path)
    return output_path


def read_values(xlsx_path, sheet_name: str = None) -> list:
    """Read a sheet back as a list of row tuples — handy for checking a
    filled template, or for reading a buyer-supplied pricing sheet you need
    to respond to."""
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return [row for row in ws.iter_rows(values_only=True)]
